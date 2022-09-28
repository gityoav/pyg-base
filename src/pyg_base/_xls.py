from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

def pd_to_excel(filename, df, 
                sheet_name = None,
                if_sheet_exists = 'replace',
                date_format = None,
                datetime_format = None, 
                na_rep = '', 
                float_format = None, 
                columns = None,
                header = True,
                index_name = None,
                mode = None,
                suffix = '.xlsx'):
    """
    Alas, df.to_excel tends to overwrite existing files completely.
    - pd_to_excel defaults to overwriting just the sheet we write
    - supports appending to sheet rather than replace
    """
    
    if index_name:
        df = df.copy()
        df.index.name = index_name
    
    fname = Path(filename).with_suffix(suffix)
    file_exists = fname.is_file()
    if mode is None:
        mode = 'a' if file_exists else 'w'
    
    wb = None
    if sheet_name is None:
        if file_exists:
            wb = load_workbook(fname)
            sheet_names = wb.sheetnames
            sheet_name = sheet_names[0]
        else:
            sheet_name = 'Sheet1'
    
    ## fake append
    if if_sheet_exists == 'append' : #fake append
        if file_exists:
            if wb is None:
                wb = load_workbook(fname)
                sheet_names = wb.sheetnames
            sheet_exists = sheet_name in sheet_names
        if sheet_exists:
            old = pd.read_excel(fname, sheet_name = sheet_name)
            df = pd.concat([old, df])
        if_sheet_exists = 'replace'
    
    with pd.ExcelWriter(fname, engine = 'openpyxl',
                        mode = mode,
                        date_format = date_format,
                        datetime_format = datetime_format,
                        if_sheet_exists = if_sheet_exists) as writer:
        df.to_excel(writer, sheet_name = sheet_name, na_rep = na_rep,
                    float_format = float_format, columns = columns, header = header)
    
    return df
        